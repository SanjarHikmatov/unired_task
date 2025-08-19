from celery import shared_task
from openpyxl import load_workbook
from .models import Card

@shared_task
def import_cards_from_excel_task(file_path):
    """
        Imports card data from an Excel file in the background using Celery.

        Parameters:
            file_path (str): The full path to the Excel file on the server.

        Returns:
            dict: A summary of the import results, including the number of successfully imported
                  records, rejected records, or an error message if the process fails.
    """
    imported, rejected = 0, 0
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [h.lower() for h in rows[0]]

        for row in rows[1:]:
            data = dict(zip(header, row))
            try:
                card_number = str(data.get("card_number", "")).strip()
                expire = str(data.get("expire", "")).strip()
                phone = str(data.get("phone", "")).strip()
                status = str(data.get("status", "")).lower()
                balance = float(data.get("balance") or 0)

                if status not in [s.value for s in Card.Status]:
                    raise ValueError("Invalid status")

                Card.objects.update_or_create(
                    card_number=card_number,
                    defaults={
                        "expire": expire,
                        "phone": phone,
                        "status": status,
                        "balance": balance,
                    },
                )
                imported += 1
            except Exception:
                rejected += 1
        return {"imported": imported, "rejected": rejected}
    except Exception as e:
        return {"error": str(e)}

